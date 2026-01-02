"""
收据打印对话框
"""
from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
                             QPushButton, QTextEdit, QMessageBox, QComboBox, QDoubleSpinBox, QScrollArea)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap
from datetime import datetime, timedelta
import os

from services.payment_service import PaymentService
from utils.printer import ReceiptPrinter
from decimal import Decimal, ROUND_HALF_UP


class ReceiptDialog(QDialog):
    """收据打印对话框"""
    
    def __init__(self, parent=None, payment_id=None):
        super().__init__(parent)
        self.payment_id = payment_id
        self.init_ui()
        self.load_receipt()
    
    def init_ui(self):
        """初始化UI"""
        self.setWindowTitle('收据打印')
        # 允许窗口可调整，设置初始大小和最小尺寸，避免内容被裁切
        self.resize(900, 700)
        self.setMinimumSize(600, 500)
        
        layout = QVBoxLayout(self)
        
        # 收据预览（使用图像预览以保证与导出的一致性）
        self.receipt_preview_label = QLabel()
        self.receipt_preview_label.setAlignment(Qt.AlignCenter)
        self.receipt_preview_label.setMinimumSize(200, 100)
        self.receipt_scroll = QScrollArea()
        self.receipt_scroll.setWidgetResizable(True)
        self.receipt_scroll.setWidget(self.receipt_preview_label)
        layout.addWidget(self.receipt_scroll)
        
        # 纸张尺寸选择
        paper_layout = QHBoxLayout()
        paper_layout.addWidget(QLabel('纸张尺寸：'))
        self.paper_size_combo = QComboBox()
        self.paper_size_combo.addItems(['收据纸 (241×93mm)'])
        self.paper_size_combo.setCurrentText('收据纸 (241×93mm)')
        self.paper_size_combo.currentTextChanged.connect(lambda x: self.load_receipt())
        paper_layout.addWidget(self.paper_size_combo)
        # 顶部偏移微调（mm）
        paper_layout.addWidget(QLabel(' 上移(mm):'))
        self.top_offset_spin = QDoubleSpinBox()
        self.top_offset_spin.setRange(-10.0, 20.0)
        self.top_offset_spin.setSingleStep(0.5)
        self.top_offset_spin.setValue(3.0)  # 安全默认，针式打印常用向上偏移
        self.top_offset_spin.valueChanged.connect(lambda _: self.load_receipt())
        paper_layout.addWidget(self.top_offset_spin)
        # 公司标题缩放系数
        paper_layout.addWidget(QLabel(' 标题缩放:'))
        self.company_scale_spin = QDoubleSpinBox()
        self.company_scale_spin.setRange(0.6, 1.2)
        self.company_scale_spin.setSingleStep(0.01)
        self.company_scale_spin.setValue(0.95)
        self.company_scale_spin.valueChanged.connect(lambda _: self.load_receipt())
        paper_layout.addWidget(self.company_scale_spin)
        paper_layout.addStretch()
        layout.addLayout(paper_layout)
        
        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        self.print_btn = QPushButton('打印')
        self.save_pdf_btn = QPushButton('保存为 PDF')
        self.save_xlsx_btn = QPushButton('保存为 Excel')
        self.close_btn = QPushButton('关闭')
        self.print_btn.clicked.connect(self.print_receipt)
        self.save_pdf_btn.clicked.connect(self.save_receipt_pdf)
        self.save_xlsx_btn.clicked.connect(self.save_receipt_xlsx)
        self.close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(self.print_btn)
        btn_layout.addWidget(self.save_pdf_btn)
        btn_layout.addWidget(self.save_xlsx_btn)
        self.save_png_btn = QPushButton('保存为图片')
        self.save_png_btn.clicked.connect(self.save_receipt_png)
        btn_layout.addWidget(self.save_png_btn)
        btn_layout.addWidget(self.close_btn)
        layout.addLayout(btn_layout)
    
    def _fmt_amount_int(self, value):
        try:
            if value is None:
                return "0"
            v = Decimal(str(value))
            return str(int(v.quantize(0, rounding=ROUND_HALF_UP)))
        except Exception:
            try:
                return str(int(round(float(value))))
            except Exception:
                return str(value)

    def _fmt_amount_two_decimals(self, value):
        """格式化金额：四舍五入到整数后显示两位小数（例如 1010 -> '1010.00'）"""
        try:
            if value is None:
                return "0.00"
            v = Decimal(str(value))
            iv = int(v.quantize(0, rounding=ROUND_HALF_UP))
            return f"{iv:.2f}"
        except Exception:
            try:
                iv = int(round(float(value)))
                return f"{iv:.2f}"
            except Exception:
                return str(value)
    
    def load_receipt(self):
        """加载收据内容"""
        try:
            payment = PaymentService.get_payment_by_id(self.payment_id)
            if not payment:
                QMessageBox.warning(self, '提示', '缴费记录不存在')
                self.reject()
                return

            # 生成与导出一致的图片预览（渲染为 PNG 再显示），保证 UI 与导出文件内容一致
            try:
                import tempfile, os
                tmp_dir = tempfile.gettempdir()
                tmp_png = os.path.join(tmp_dir, f"receipt_preview_{self.payment_id}.png")
                paper_size = self.paper_size_combo.currentText()
                top_offset = float(getattr(self, 'top_offset_spin', None).value()) if getattr(self, 'top_offset_spin', None) else 0.0
                comp_scale = float(getattr(self, 'company_scale_spin', None).value()) if getattr(self, 'company_scale_spin', None) else 1.0
                from utils.printer import ReceiptPrinter
                printer = ReceiptPrinter(paper_size=paper_size, top_offset_mm=top_offset, company_font_scale_adj=comp_scale)
                # 预览使用 300dpi 输出与导出匹配像素比，UI 会缩放显示
                ok = printer.render_receipt_to_image(self.payment_id, tmp_png, dpi=300)
                if ok and os.path.exists(tmp_png):
                    pix = QPixmap(tmp_png)
                    if not pix.isNull():
                        # 将图片按滚动视口宽度缩放以便预览，但保持比例
                        vw = self.receipt_scroll.viewport().width() or 800
                        scaled = pix.scaledToWidth(vw - 20, Qt.SmoothTransformation)
                        self.receipt_preview_label.setPixmap(scaled)
                        self.receipt_preview_label.resize(scaled.size())
                        return
            except Exception:
                # 渲染失败则回退到 HTML 预览（保持原有体验）
                pass

            # 回退：生成 HTML 预览（当图片渲染不可用时）
            receipt_html = self.generate_receipt_html(payment)
            # 如果回退为 HTML，需要把 label 显示 HTML（转换为 QPixmap via QTextDocument）
            try:
                from PyQt5.QtGui import QTextDocument
                doc = QTextDocument()
                doc.setHtml(receipt_html)
                img = QPixmap(doc.size().toSize())
                img.fill(Qt.white)
                painter = QPainter(img)
                doc.drawContents(painter)
                painter.end()
                self.receipt_preview_label.setPixmap(img)
                self.receipt_preview_label.resize(img.size())
            except Exception:
                # 最后回退：直接把 HTML 放到文本框样式字符串显示为简单文本
                self.receipt_preview_label.setText(receipt_html)
        except Exception as e:
            QMessageBox.critical(self, '错误', f'加载收据失败：{str(e)}')
            self.reject()
    
    def generate_receipt_text(self, payment):
        """生成收据文本"""
        # 保留原方法签名，但生成 HTML 格式字符串用于预览
        # 使用与打印一致的样式：公司抬头、表格、合计大写/小写、签名行
        try:
            from utils.printer import ReceiptPrinter
            total_amount = float(payment.amount) if payment.amount else 0.0
            paid_amount = float(payment.paid_amount) if getattr(payment, 'paid_amount', None) else 0.0
            # 显示实际金额：优先显示已缴金额（如果有），否则显示账单总额
            display_amount = paid_amount if paid_amount > 0 else total_amount
            # 合计大写使用整数元金额
            try:
                display_amount_int = int(Decimal(str(display_amount)).quantize(0, rounding=ROUND_HALF_UP))
            except Exception:
                display_amount_int = int(round(float(display_amount)))
            upper_amount = ReceiptPrinter()._num_to_rmb_upper(display_amount_int)
        except Exception:
            total_amount = float(payment.amount) if payment.amount else 0.0
            display_amount = total_amount
            upper_amount = ""

        # 计费起止展示
        billing_period = ""
        if payment.billing_start_date and payment.billing_end_date:
            try:
                end_display = payment.billing_end_date - timedelta(days=1)
            except Exception:
                end_display = payment.billing_end_date
            billing_period = f"{payment.billing_start_date.strftime('%Y.%m.%d')}–{end_display.strftime('%Y.%m.%d')}"
        # 使用最近一笔流水（如果存在）来显示“本次实收”周期与金额；否则回退到累计显示
        paid_period = ""
        paid_amount_display = ""
        try:
            from services.payment_transaction_service import PaymentTransactionService
            last_tx = PaymentTransactionService.get_last_transaction(payment.id)
            if last_tx:
                # 计算本次流水对应的月份数（按月均摊的月额近似计算）
                if payment.billing_months and payment.billing_months > 0:
                    monthly_amount = float(payment.amount) / payment.billing_months
                    # 避免除以0
                    months_in_tx = int(round(float(last_tx.amount) / monthly_amount)) if monthly_amount > 0 else 0
                else:
                    months_in_tx = 0

                def add_months(dt, months):
                    month = dt.month - 1 + months
                    year = dt.year + month // 12
                    month = month % 12 + 1
                    import calendar
                    day = min(dt.day, calendar.monthrange(year, month)[1])
                    return dt.replace(year=year, month=month, day=day)

                # 计算本次流水的开始日期（基于已累计 paid_months）
                total_paid = int(payment.paid_months) if payment.paid_months else 0
                prev_paid = max(0, total_paid - months_in_tx)
                start_tx = add_months(payment.billing_start_date, prev_paid)
                end_tx = add_months(start_tx, months_in_tx)
                try:
                    end_tx_display = end_tx - timedelta(days=1)
                except Exception:
                    end_tx_display = end_tx
                paid_period = f"{start_tx.strftime('%Y.%m.%d')}–{end_tx_display.strftime('%Y.%m.%d')}"
                paid_amount_display = self._fmt_amount_two_decimals(last_tx.amount)
            else:
                # 无流水则退回到累计已缴金额显示
                if getattr(payment, 'paid_months', 0) and payment.paid_months > 0 and payment.billing_start_date:
                    def add_months(dt, months):
                        month = dt.month - 1 + months
                        year = dt.year + month // 12
                        month = month % 12 + 1
                        import calendar
                        day = min(dt.day, calendar.monthrange(year, month)[1])
                        return dt.replace(year=year, month=month, day=day)
                    start = payment.billing_start_date
                    end_paid = add_months(start, int(payment.paid_months))
                    try:
                        end_paid_display = end_paid - timedelta(days=1)
                    except Exception:
                        end_paid_display = end_paid
                    paid_period = f"{start.strftime('%Y.%m.%d')}–{end_paid_display.strftime('%Y.%m.%d')}"
                    paid_amount_display = self._fmt_amount_two_decimals(payment.paid_amount) if payment.paid_amount else ""
        except Exception:
            paid_period = ""
            paid_amount_display = ""

        # 获取纸张设置以调整预览样式（确保预览长宽比与导出图片一致）
        curr_paper = self.paper_size_combo.currentText()
        is_wide = "241" in curr_paper
        is_narrow = "80" in curr_paper

        # 计算目标纸张物理尺寸（mm）
        if curr_paper == '收据纸 (241×93mm)':
            target_w_mm, target_h_mm = 241.0, 93.0
        elif curr_paper == '收据纸 (80×200mm)':
            target_w_mm, target_h_mm = 80.0, 200.0
        else:
            target_w_mm, target_h_mm = 210.0, 297.0

        # 预览宽度使用预览控件当前可用宽度，保证显示比例与导出图片一致
        try:
            preview_widget_width = max(300, self.receipt_preview.viewport().width() - 20)
        except Exception:
            preview_widget_width = 500
        # 计算高度以匹配纸张的长宽比
        try:
            preview_height = int(preview_widget_width * (target_h_mm / target_w_mm))
        except Exception:
            preview_height = int(preview_widget_width * (93.0 / 241.0)) if is_wide else int(preview_widget_width * (297.0 / 210.0))

        paper_width = f"{preview_widget_width}px"
        paper_height = f"{preview_height}px"

        # 基础样式值（可随纸张类型调整）
        padding = "30px 40px"
        empty_rows_count = 6
        base_font_size = "14px"
        title_size = "20px"
        company_size = "22px"
        cell_padding = "8px"
        margin_bottom_divider = "12px"

        if is_wide:
            # 宽纸 241x93mm 专属紧凑样式
            padding = "15px 20px"
            empty_rows_count = 0
            base_font_size = "12px"
            title_size = "16px"
            company_size = "18px"
            cell_padding = "4px"
            margin_bottom_divider = "5px"
        elif is_narrow:
            padding = "10px 5px"
            empty_rows_count = 2
            base_font_size = "12px"
            title_size = "16px"
            company_size = "16px"
            cell_padding = "4px"
            margin_bottom_divider = "8px"

        # 准备 LOGO HTML
        # ... (保持原有 LOGO 逻辑) ...
        # (略去重复代码，此处假设 tool 会保留上下文，但为了安全我通常包含上下文)
        # 上下文太长，只替换参数定义部分和 HTML header

        logo_html = ""
        try:
            logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'logo.jpg')
            if os.path.exists(logo_path):
                logo_url = "file:///" + os.path.abspath(logo_path).replace("\\", "/")
                # Logo高度也自适应
                logo_h = "40" if is_wide else "50"
                logo_html = f'<img src="{logo_url}" height="{logo_h}" style="vertical-align:middle; margin-right:10px;" />'
        except Exception:
            pass

        # 票据序号：前缀保留缴费创建日期（YYYYMMDD），后三位为当日打印序号（预览使用当日下一个序号，不写流水）
        try:
            payment_date_str = payment.created_at.strftime('%Y%m%d') if getattr(payment, 'created_at', None) else datetime.now().strftime('%Y%m%d')
            from services.print_service import PrintService
            payment_seq = PrintService.get_today_sequence()
        except Exception:
            payment_date_str = datetime.now().strftime('%Y%m%d')
            payment_seq = getattr(payment, 'id', 1) % 1000 if getattr(payment, 'id', None) else 1

        html = f"""
<html>
<head>
<meta charset="utf-8" />
<style>
body {{ 
    font-family: "宋体", "Microsoft YaHei", Arial; 
    font-size: {base_font_size}; 
    background: #f0f0f0;
    margin: 0;
    padding: 20px;
    display: flex; justify-content: center; /* 居中显示 */
}}
.receipt-paper {{
    background: #fff;
    width: {paper_width};
    padding: {padding};
    box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    border: 1px solid #ddd;
    /* 强制盒模型，避免 padding 撑大 */
    box-sizing: border-box; 
}}
.company {{ font-size:{company_size}; font-weight:700; text-align:center; color:#333; }}
.divider {{ display:none; }}
.title {{ font-size:{title_size}; font-weight:700; text-align:center; margin-top:2px; margin-bottom:5px; }}
.receipt-no {{ text-align:right; font-size:{base_font_size}; color:#666; margin-bottom:2px; }}
.info {{ margin-top:2px; margin-bottom:5px; font-size:{base_font_size}; }}
table {{ border-collapse:collapse; width:100%; }}
td, th {{ border:1px solid #333; padding:{cell_padding}; vertical-align:middle; font-size:{base_font_size}; }}
th {{ background:#f5f5f5; font-weight:600; }}
.right {{ text-align:right; }}
.center {{ text-align:center; }}
.note {{ margin-top:5px; font-size:{base_font_size}; color:#666; }}
.signature {{ margin-top:15px; font-size:{base_font_size}; }}
</style>
</head>
<body>
  <div class="receipt-paper" style="width:{paper_width}; height:{paper_height};">
  <div class="company">{logo_html}四川盛涵物业服务有限公司</div>
  <div class="title">收费收据</div>
  <div class="receipt-no">NO:{payment_date_str}{payment_seq:03d}</div>
  <div class="info">
    <table style="border:none; width:100%;">
      <tr style="border:none;">
        <td style="border:none; padding:2px;">户名：{payment.resident.name}</td>
        <td style="border:none; padding:2px;">房号：{getattr(payment.resident, 'full_room_no', payment.resident.room_no)}</td>
                <td style="border:none; padding:2px;">日期：{datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}</td>
        <td style="border:none; padding:2px; text-align:right;">NO:{payment_date_str}{payment_seq:03d}</td>
      </tr>
    </table>
  </div>

  <table style="width:60%; margin:0 auto; border:1px solid #000; border-collapse:collapse;">
    <tr>
      <th style="width:25%; text-align:center;">收费项目</th>
      <th style="width:45%; text-align:center;">起止时间</th>
      <th style="width:15%; text-align:center;">金额（元）</th>
      <th style="width:15%; text-align:center;">备注</th>
    </tr>
    <tr>
      <td>{payment.charge_item.name if payment.charge_item else ""}</td>
      <td class="center">{billing_period}</td>
      <td class="center">{self._fmt_amount_two_decimals(total_amount)}</td>
      <td></td>
    </tr>
""" 
        # 预留多行空行以匹配纸质样式
        for i in range(empty_rows_count):
            html += "    <tr><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td></tr>\n"
        # 如果有实付信息，插入实付周期和实付金额行（作为合并单元格）
        if paid_period:
            html += f"""    <tr>
      <td colspan="2" style="border:1px solid #000; padding-left:8px;">实收周期</td>
      <td colspan="2" style="border:1px solid #000; text-align:center;">{paid_period}</td>
    </tr>
    <tr>
      <td colspan="2" style="border:1px solid #000; padding-left:8px;">实收金额（元）</td>
      <td colspan="2" style="border:1px solid #000; text-align:center;">{paid_amount_display}</td>
    </tr>
"""

        # 合计行，放入同一 table 以保证被表格包裹
        html += f"""    <tr>
      <td style="font-weight:700; border:1px solid #000; padding-left:8px;">合计金额大写</td>
      <td style="border:1px solid #000;">{upper_amount}</td>
            <td style="font-weight:700; border:1px solid #000; text-align:center;">合计金额小写</td>
      <td style="border:1px solid #000; text-align:center;">{self._fmt_amount_two_decimals(display_amount)}元</td>
    </tr>
    <tr>
      <td colspan="4" style="border:1px solid #000; padding:6px 8px; text-align:left;">请确认您的缴费金额，如有疑问请咨询物业服务中心</td>
    </tr>
  </table>

  <div class="signature">
    <span style="display:inline-block; width:50%;">收款人：__________</span>
    <span style="display:inline-block; width:49%;">收款单位盖章：</span>
  </div>
</div>

</body>
</html>
"""
        return html

    def generate_receipt_html(self, payment):
        """兼容旧接口：返回 HTML 格式的收据预览"""
        return self.generate_receipt_text(payment)
    
    def print_receipt(self):
        """打印收据"""
        try:
            paper_size = self.paper_size_combo.currentText()
            top_offset = float(getattr(self, 'top_offset_spin', None).value()) if getattr(self, 'top_offset_spin', None) else 0.0
            comp_scale = float(getattr(self, 'company_scale_spin', None).value()) if getattr(self, 'company_scale_spin', None) else 1.0
            printer = ReceiptPrinter(paper_size=paper_size, top_offset_mm=top_offset, company_font_scale_adj=comp_scale)
            if printer.print_receipt(self.payment_id):
                QMessageBox.information(self, '成功', '打印成功')
            else:
                QMessageBox.warning(self, '提示', '打印取消或失败')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'打印失败：{str(e)}')

    def save_receipt_pdf(self):
        """将收据保存为 PDF 到 exports 目录"""
        try:
            import os
            from PyQt5.QtWidgets import QFileDialog

            # 推荐让用户选择保存位置，便于诊断权限问题
            default_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'exports'))
            if not os.path.exists(default_dir):
                try:
                    os.makedirs(default_dir, exist_ok=True)
                except Exception as e:
                    # 如果创建失败，回退到当前工作目录
                    default_dir = os.getcwd()

            suggested = os.path.join(default_dir, f"receipt_{self.payment_id:06d}.pdf")
            path, _ = QFileDialog.getSaveFileName(self, "保存为 PDF", suggested, "PDF Files (*.pdf)")
            if not path:
                QMessageBox.information(self, '提示', '已取消保存')
                return

            # 确保后缀为 .pdf
            if not path.lower().endswith('.pdf'):
                path = path + '.pdf'

            paper_size = self.paper_size_combo.currentText()
            top_offset = float(getattr(self, 'top_offset_spin', None).value()) if getattr(self, 'top_offset_spin', None) else 0.0
            comp_scale = float(getattr(self, 'company_scale_spin', None).value()) if getattr(self, 'company_scale_spin', None) else 1.0
            printer = ReceiptPrinter(paper_size=paper_size, top_offset_mm=top_offset, company_font_scale_adj=comp_scale)
            success = printer.print_receipt(self.payment_id, output_file=path)
            if success:
                QMessageBox.information(self, '成功', f'已保存 PDF：{path}')
            else:
                # 打印器内部会打印异常到控制台，给用户更详细的提示
                QMessageBox.warning(self, '提示', '保存 PDF 失败（请检查控制台日志以获取详细错误）')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'保存 PDF 失败：{str(e)}')

    def save_receipt_xlsx(self):
        """将当前收据导出为 Excel (.xlsx) 文件"""
        try:
            import os
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            from PyQt5.QtWidgets import QFileDialog

            # 准备目录与默认文件名
            default_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'exports'))
            if not os.path.exists(default_dir):
                try:
                    os.makedirs(default_dir, exist_ok=True)
                except Exception:
                    default_dir = os.getcwd()
            suggested = os.path.join(default_dir, f"receipt_{self.payment_id:06d}.xlsx")
            path, _ = QFileDialog.getSaveFileName(self, "保存为 Excel", suggested, "Excel Files (*.xlsx)")
            if not path:
                QMessageBox.information(self, '提示', '已取消保存')
                return
            if not path.lower().endswith('.xlsx'):
                path = path + '.xlsx'

            # 获取 payment 数据
            payment = PaymentService.get_payment_by_id(self.payment_id)
            if not payment:
                QMessageBox.warning(self, '提示', '缴费记录不存在')
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "收据"

            # 样式
            bold = Font(bold=True)
            center = Alignment(horizontal='center', vertical='center')
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # 标题
            ws.merge_cells('A1:D1')
            ws['A1'] = "四川盛涵物业服务有限公司"
            ws['A1'].font = Font(size=18, bold=True)
            ws['A1'].alignment = center
            ws.merge_cells('A2:D2')
            ws['A2'] = "收费收据"
            ws['A2'].font = Font(size=14, bold=True)
            ws['A2'].alignment = center

            # 基本信息
            ws['A4'] = f"户名：{payment.resident.name}"
            ws['B4'] = f"房号：{getattr(payment.resident, 'full_room_no', payment.resident.room_no)}"
            ws['C4'] = f"日期：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            # 表头
            start_row = 6
            headers = ['收费项目', '起止时间', '金额（元）', '备注']
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col, value=h)
                cell.font = bold
                cell.alignment = center
                cell.border = border

            # 明细行
            billing_period = ""
            if payment.billing_start_date and payment.billing_end_date:
                try:
                    end_display = payment.billing_end_date - timedelta(days=1)
                except Exception:
                    end_display = payment.billing_end_date
                billing_period = f"{payment.billing_start_date.strftime('%Y.%m.%d')}–{end_display.strftime('%Y.%m.%d')}"
            detail_row = start_row + 1
            ws.cell(row=detail_row, column=1, value=payment.charge_item.name if payment.charge_item else "")
            ws.cell(row=detail_row, column=2, value=billing_period)
            # 写为整数元
            try:
                amt_int = int(Decimal(str(payment.amount)).quantize(0, rounding=ROUND_HALF_UP))
            except Exception:
                try:
                    amt_int = int(round(float(payment.amount)))
                except Exception:
                    amt_int = int(round(float(payment.amount))) if payment.amount else 0
            ws.cell(row=detail_row, column=3, value=f"{amt_int:.2f}")
            ws.cell(row=detail_row, column=4, value="")
            for col in range(1, 5):
                c = ws.cell(row=detail_row, column=col)
                c.border = border
                c.alignment = center

            # 预留空行
            for r in range(detail_row + 1, detail_row + 1 + 6):
                for col in range(1, 5):
                    c = ws.cell(row=r, column=col, value="")
                    c.border = border

            # 合计行
            total_row = detail_row + 1 + 6
            ws.cell(row=total_row, column=1, value="合计金额大写").font = bold
            ws.cell(row=total_row, column=2, value=self._get_upper_amount_cell(payment)).alignment = center
            ws.cell(row=total_row, column=3, value="合计金额小写").font = bold
            try:
                total_int = int(Decimal(str(payment.amount)).quantize(0, rounding=ROUND_HALF_UP))
            except Exception:
                try:
                    total_int = int(round(float(payment.amount)))
                except Exception:
                    total_int = int(round(float(payment.amount))) if payment.amount else 0
            ws.cell(row=total_row, column=4, value=f"{total_int:.2f}元")
            for col in range(1, 5):
                c = ws.cell(row=total_row, column=col)
                c.border = border
                c.alignment = center

            # 提示行（跨列）
            tip_row = total_row + 1
            ws.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row, end_column=4)
            ws.cell(row=tip_row, column=1, value="请确认您的缴费金额，如有疑问请咨询物业服务中心").alignment = Alignment(horizontal='left')
            ws.cell(row=tip_row, column=1).border = border

            # 列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12

            wb.save(path)
            QMessageBox.information(self, '成功', f'已保存 Excel：{path}')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'保存 Excel 失败：{str(e)}')

    def _get_upper_amount_cell(self, payment):
        """辅助：返回中文大写金额字符串（与打印器保持一致）"""
        try:
            from utils.printer import ReceiptPrinter
            return ReceiptPrinter()._num_to_rmb_upper(float(payment.amount))
        except Exception:
            return ""

    def save_receipt_png(self):
        """将收据渲染为 PNG 图像保存（备用导出，便于排版确认）"""
        try:
            import os
            from PyQt5.QtWidgets import QFileDialog

            default_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'exports'))
            if not os.path.exists(default_dir):
                try:
                    os.makedirs(default_dir, exist_ok=True)
                except Exception:
                    default_dir = os.getcwd()

            suggested = os.path.join(default_dir, f"receipt_{self.payment_id:06d}.png")
            path, _ = QFileDialog.getSaveFileName(self, "保存为图片", suggested, "PNG Files (*.png);;All Files (*)")
            if not path:
                QMessageBox.information(self, '提示', '已取消保存')
                return
            if not path.lower().endswith('.png'):
                path = path + '.png'

            paper_size = self.paper_size_combo.currentText()
            top_offset = float(getattr(self, 'top_offset_spin', None).value()) if getattr(self, 'top_offset_spin', None) else 0.0
            comp_scale = float(getattr(self, 'company_scale_spin', None).value()) if getattr(self, 'company_scale_spin', None) else 1.0
            printer = ReceiptPrinter(paper_size=paper_size, top_offset_mm=top_offset, company_font_scale_adj=comp_scale)
            ok = printer.render_receipt_to_image(self.payment_id, path, dpi=300)
            if ok:
                QMessageBox.information(self, '成功', f'已保存图片：{path}')
            else:
                QMessageBox.warning(self, '提示', '保存图片失败（请检查控制台日志）')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'保存图片失败：{str(e)}')


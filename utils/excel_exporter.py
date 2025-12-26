"""
Excel导出工具
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from services.payment_service import PaymentService
from services.resident_service import ResidentService


class ExcelExporter:
    """Excel导出工具类"""
    
    @staticmethod
    def export_unpaid_list(period, file_path):
        """导出欠费清单到Excel
        
        Args:
            period: 缴费周期（格式：YYYY-MM）
            file_path: 保存路径
            
        Returns:
            bool: 是否成功
        """
        try:
            # 获取欠费记录
            unpaid_payments = PaymentService.get_unpaid_payments_by_period(period)
            
            # 创建Excel工作簿
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"欠费清单_{period}"
            
            # 设置标题样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            # 标题行
            headers = ['房号', '姓名', '电话', '收费项目', '计费周期', '总月数', '已缴月数', '总金额', '已缴金额', '欠费金额', '生成时间']
            sheet.append(headers)
            
            # 设置标题样式
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 数据行
            total_unpaid = 0.0
            for payment in unpaid_payments:
                unpaid_amount = float(payment.amount) - (float(payment.paid_amount) if payment.paid_amount else 0.0)
                total_unpaid += unpaid_amount
                
                row = [
                    payment.resident.room_no,
                    payment.resident.name,
                    payment.resident.phone or '',
                    payment.charge_item.name,
                    f"{payment.billing_start_date.strftime('%Y-%m-%d')} 至 {payment.billing_end_date.strftime('%Y-%m-%d')}" if payment.billing_start_date and payment.billing_end_date else payment.period,
                    payment.billing_months,
                    payment.paid_months,
                    float(payment.amount),
                    float(payment.paid_amount) if payment.paid_amount else 0.0,
                    unpaid_amount,
                    payment.created_at.strftime('%Y-%m-%d %H:%M:%S') if payment.created_at else ''
                ]
                sheet.append(row)
            
            # 添加汇总行
            sheet.append([])
            summary_row = ['', '', '', '', '', '', '', '合计', '', f'¥{total_unpaid:.2f}', '']
            sheet.append(summary_row)
            
            # 设置汇总行样式
            summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            summary_font = Font(bold=True, size=11)
            for col_idx in range(1, len(headers) + 1):
                cell = sheet.cell(row=sheet.max_row, column=col_idx)
                cell.fill = summary_fill
                cell.font = summary_font
            
            # 设置列宽
            column_widths = [12, 12, 15, 20, 30, 10, 10, 12, 12, 12, 20]
            for col_idx, width in enumerate(column_widths, start=1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
            
            # 设置数据格式
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row-2):
                # 金额列右对齐
                for col_idx in [8, 9, 10]:  # 总金额、已缴金额、欠费金额
                    if col_idx <= len(row):
                        row[col_idx-1].alignment = Alignment(horizontal='right')
            
            # 保存文件
            workbook.save(file_path)
            return True
            
        except Exception as e:
            raise Exception(f"导出失败：{str(e)}")
    
    @staticmethod
    def export_payments(period=None, file_path=None):
        """导出缴费记录到Excel
        
        Args:
            period: 缴费周期（可选）
            file_path: 保存路径
        """
        try:
            if period:
                payments = PaymentService.get_payments_by_period(period)
            else:
                payments = PaymentService.get_all_payments()
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"缴费记录_{period or '全部'}"
            
            # 标题样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            headers = ['房号', '姓名', '收费项目', '缴费周期', '计费周期', '总月数', '已缴月数', '总金额', '已缴金额', '缴费状态', '缴费时间']
            sheet.append(headers)
            
            # 设置标题样式
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 数据行
            for payment in payments:
                if payment.paid == 1:
                    status = f"已缴费 ({payment.paid_months}/{payment.billing_months}个月)"
                elif payment.paid_months > 0:
                    status = f"部分缴费 ({payment.paid_months}/{payment.billing_months}个月)"
                else:
                    status = "未缴费"
                
                billing_period = f"{payment.billing_start_date.strftime('%Y-%m-%d')} 至 {payment.billing_end_date.strftime('%Y-%m-%d')}" if payment.billing_start_date and payment.billing_end_date else payment.period
                
                row = [
                    payment.resident.room_no,
                    payment.resident.name,
                    payment.charge_item.name,
                    payment.period,
                    billing_period,
                    payment.billing_months,
                    payment.paid_months,
                    float(payment.amount),
                    float(payment.paid_amount) if payment.paid_amount else 0.0,
                    status,
                    payment.paid_time.strftime('%Y-%m-%d %H:%M:%S') if payment.paid_time else ''
                ]
                sheet.append(row)
            
            # 设置列宽
            column_widths = [12, 12, 20, 12, 30, 10, 10, 12, 12, 20, 20]
            for col_idx, width in enumerate(column_widths, start=1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
            
            workbook.save(file_path)
            return True
            
        except Exception as e:
            raise Exception(f"导出失败：{str(e)}")


"""
报表生成工具
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime
from services.payment_service import PaymentService
from services.charge_service import ChargeService


class ReportGenerator:
    """报表生成工具类"""
    
    @staticmethod
    def generate_monthly_report(period, file_path):
        """生成月度收费统计报表
        
        Args:
            period: 缴费周期（格式：YYYY-MM）
            file_path: 保存路径
        """
        try:
            # 获取统计数据
            stats = PaymentService.get_statistics_by_period(period)
            payments = PaymentService.get_payments_by_period(period)
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"月度统计_{period}"
            
            # 标题样式
            title_font = Font(bold=True, size=16)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            # 报表标题
            sheet.merge_cells('A1:D1')
            title_cell = sheet['A1']
            title_cell.value = f"{period} 月度收费统计报表"
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center')
            
            sheet.append([])
            
            # 统计摘要
            summary_row = 3
            sheet.cell(row=summary_row, column=1).value = "统计摘要"
            sheet.cell(row=summary_row, column=1).font = Font(bold=True, size=14)
            
            summary_data = [
                ['总账单数', stats['total_count']],
                ['已缴费数', stats['paid_count']],
                ['未缴费数', stats['unpaid_count']],
                ['总金额', f"¥{stats['total_amount']:.2f}"],
                ['已缴费金额', f"¥{stats['paid_amount']:.2f}"],
                ['欠费金额', f"¥{stats['unpaid_amount']:.2f}"],
                ['缴费率', f"{(stats['paid_count']/stats['total_count']*100) if stats['total_count'] > 0 else 0:.1f}%"]
            ]
            
            for idx, (label, value) in enumerate(summary_data, start=summary_row + 1):
                sheet.cell(row=idx, column=1).value = label
                sheet.cell(row=idx, column=2).value = value
                sheet.cell(row=idx, column=1).font = Font(bold=True)
            
            sheet.append([])
            
            # 按收费项目统计
            charge_item_stats = {}
            for payment in payments:
                item_name = payment.charge_item.name
                if item_name not in charge_item_stats:
                    charge_item_stats[item_name] = {
                        'total': 0.0,
                        'paid': 0.0,
                        'count': 0,
                        'paid_count': 0
                    }
                charge_item_stats[item_name]['total'] += float(payment.amount)
                charge_item_stats[item_name]['count'] += 1
                if payment.paid == 1:
                    charge_item_stats[item_name]['paid'] += float(payment.amount)
                    charge_item_stats[item_name]['paid_count'] += 1
                elif payment.paid_amount:
                    charge_item_stats[item_name]['paid'] += float(payment.paid_amount)
            
            # 收费项目明细表
            detail_row = sheet.max_row + 2
            sheet.cell(row=detail_row, column=1).value = "收费项目明细"
            sheet.cell(row=detail_row, column=1).font = Font(bold=True, size=14)
            
            detail_row += 1
            headers = ['收费项目', '账单数', '已缴费数', '总金额', '已缴费金额', '欠费金额']
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=detail_row, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for item_name, item_stats in charge_item_stats.items():
                detail_row += 1
                unpaid = item_stats['total'] - item_stats['paid']
                row = [
                    item_name,
                    item_stats['count'],
                    item_stats['paid_count'],
                    f"¥{item_stats['total']:.2f}",
                    f"¥{item_stats['paid']:.2f}",
                    f"¥{unpaid:.2f}"
                ]
                for col_idx, value in enumerate(row, start=1):
                    sheet.cell(row=detail_row, column=col_idx).value = value
            
            # 设置列宽
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 12
            sheet.column_dimensions['C'].width = 12
            sheet.column_dimensions['D'].width = 15
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 15
            
            # 添加生成时间
            sheet.append([])
            sheet.cell(row=sheet.max_row, column=1).value = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            workbook.save(file_path)
            return True
            
        except Exception as e:
            raise Exception(f"生成报表失败：{str(e)}")

    @staticmethod
    def generate_daily_report(period, file_path):
        """生成日度收费统计报表（按 day 聚合），period 格式 YYYY-MM"""
        try:
            # 获取当月的缴费记录
            payments = PaymentService.get_payments_by_period(period)

            # 统计摘要（与月度相同的字段）
            total_count = len(payments)
            paid_count = sum(1 for p in payments if p.paid == 1)
            unpaid_count = total_count - paid_count
            total_amount = sum(float(p.amount) for p in payments) if payments else 0.0
            paid_amount = sum(float(p.paid_amount) for p in payments if p.paid_amount) if payments else 0.0
            unpaid_amount = total_amount - paid_amount

            # 按天聚合（使用 billing_start_date 的日作为分组）
            day_stats = {}
            # 以及按收费项目聚合（用于明细）
            charge_item_stats = {}
            for p in payments:
                # day grouping
                if p.billing_start_date:
                    day = p.billing_start_date.day
                else:
                    day = 1
                if day not in day_stats:
                    day_stats[day] = {'total': 0.0, 'paid': 0.0, 'count': 0}
                day_stats[day]['total'] += float(p.amount)
                day_stats[day]['paid'] += float(p.paid_amount) if p.paid_amount else 0.0
                day_stats[day]['count'] += 1

                # charge item detail
                item_name = p.charge_item.name if p.charge_item else '未知'
                if item_name not in charge_item_stats:
                    charge_item_stats[item_name] = {'total': 0.0, 'paid': 0.0, 'count': 0, 'paid_count': 0}
                charge_item_stats[item_name]['total'] += float(p.amount)
                charge_item_stats[item_name]['paid'] += float(p.paid_amount) if p.paid_amount else 0.0
                charge_item_stats[item_name]['count'] += 1
                if p.paid == 1:
                    charge_item_stats[item_name]['paid_count'] += 1

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"日度统计_{period}"

            # 标题
            title_font = Font(bold=True, size=16)
            sheet.merge_cells('A1:E1')
            sheet['A1'] = f"{period} 日度收费统计报表"
            sheet['A1'].font = title_font
            sheet['A1'].alignment = Alignment(horizontal='center')

            sheet.append([])
            # 统计摘要（左侧）
            summary_row = sheet.max_row + 1
            sheet.cell(row=summary_row, column=1).value = "统计摘要"
            sheet.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
            summary_data = [
                ('总账单数', total_count),
                ('已缴费数', paid_count),
                ('未缴费数', unpaid_count),
                ('总金额', f"¥{total_amount:.2f}"),
                ('已缴费金额', f"¥{paid_amount:.2f}"),
                ('欠费金额', f"¥{unpaid_amount:.2f}"),
                ('缴费率', f"{(paid_count/total_count*100) if total_count>0 else 0:.1f}%")
            ]
            for i, (label, value) in enumerate(summary_data, start=summary_row + 1):
                sheet.cell(row=i, column=1).value = label
                sheet.cell(row=i, column=2).value = value

            sheet.append([])
            # 日汇总表头
            sheet.append(['日期', '账单数', '日合计(¥)', '已缴(¥)', '欠费(¥)'])
            for day in sorted(day_stats.keys()):
                d = day_stats[day]
                unpaid_day = d['total'] - d['paid']
                sheet.append([f"{period}-{day:02d}", d['count'], f"¥{d['total']:.2f}", f"¥{d['paid']:.2f}", f"¥{unpaid_day:.2f}"])

            sheet.append([])
            # 收费项目明细（账单数/已缴费数/总金额/已缴金额/欠费金额）
            sheet.append(['收费项目', '账单数', '已缴费数', '总金额', '已缴金额', '欠费金额'])
            for name, stats_item in charge_item_stats.items():
                unpaid_item = stats_item['total'] - stats_item['paid']
                sheet.append([name, stats_item['count'], stats_item['paid_count'], f"¥{stats_item['total']:.2f}", f"¥{stats_item['paid']:.2f}", f"¥{unpaid_item:.2f}"])

            # 列宽
            sheet.column_dimensions['A'].width = 18
            sheet.column_dimensions['B'].width = 12
            sheet.column_dimensions['C'].width = 18
            sheet.column_dimensions['D'].width = 18
            sheet.column_dimensions['E'].width = 18
            sheet.column_dimensions['F'].width = 18

            sheet.append([])
            sheet.cell(row=sheet.max_row, column=1).value = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            workbook.save(file_path)
            return True

        except Exception as e:
            raise Exception(f"生成日度报表失败：{str(e)}")

    @staticmethod
    def generate_year_report(year, file_path):
        """生成年度收费统计报表（按月聚合并按收费项目汇总）"""
        try:
            # 获取年度内的缴费记录
            stats = PaymentService.get_statistics_by_year(int(year))
            # stats: {'year':year,'total_amount':..., 'by_item':[(name,amt)...]}

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"年度统计_{year}"

            title_font = Font(bold=True, size=16)
            sheet.merge_cells('A1:D1')
            sheet['A1'] = f"{year} 年度收费统计报表"
            sheet['A1'].font = title_font
            sheet['A1'].alignment = Alignment(horizontal='center')

            sheet.append([])
            sheet.append(['统计项', '数值'])
            sheet.append(['年度账单总额', f"¥{stats.get('total_amount', 0.0):.2f}"])
            sheet.append(['已缴金额', f"¥{stats.get('paid_amount', 0.0):.2f}"])
            sheet.append(['欠费金额', f"¥{stats.get('unpaid_amount', 0.0):.2f}"])
            sheet.append([])

            # 按收费项目明细：展示年度合计、已缴、欠费
            sheet.append(['收费项目', '年度合计(¥)', '已缴(¥)', '欠费(¥)'])
            for entry in stats.get('by_item', []):
                # entry = (name, total, paid, unpaid)
                if len(entry) == 4:
                    name, total, paid, unpaid = entry
                else:
                    name = entry[0]
                    total = entry[1] if len(entry) > 1 else 0.0
                    paid = entry[2] if len(entry) > 2 else 0.0
                    unpaid = total - paid
                sheet.append([name, f"¥{total:.2f}", f"¥{paid:.2f}", f"¥{unpaid:.2f}"])

            sheet.append([])
            sheet.cell(row=sheet.max_row, column=1).value = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            workbook.save(file_path)
            return True
        except Exception as e:
            raise Exception(f"生成年度报表失败：{str(e)}")


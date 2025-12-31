"""
Excel导入工具
"""
import openpyxl
from datetime import datetime
from PyQt5.QtWidgets import QMessageBox
from services.resident_service import ResidentService


class ExcelImporter:
    """Excel导入工具类"""
    
    @staticmethod
    def import_residents(file_path):
        """从Excel文件导入住户信息
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            tuple: (成功数量, 失败数量, 错误列表)
        """
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            success_count = 0
            fail_count = 0
            errors = []
            
            # 跳过标题行，从第2行开始
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # 跳过空行
                if not row[0] if row else None:
                    continue
                
                try:
                    # 解析数据（列顺序：楼栋、单元、房号、姓名、电话、面积、入住日期[, 身份(identity), 房屋类型(property_type)]）
                    building = str(row[0]).strip() if row[0] else ''
                    unit = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    room_no = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                    name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                    phone = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                    area = float(row[5]) if len(row) > 5 and row[5] else 0.0
                    
                    # 解析入住日期
                    move_in_date = None
                    if len(row) > 4 and row[4]:
                        if isinstance(row[4], datetime):
                            move_in_date = row[4]
                        elif isinstance(row[4], str):
                            try:
                                move_in_date = datetime.strptime(row[4], '%Y-%m-%d')
                            except:
                                try:
                                    move_in_date = datetime.strptime(row[4], '%Y/%m/%d')
                                except:
                                    pass
                    
                    # 可选字段：入住日期后为 identity, property_type （可选）
                    identity = str(row[6]).strip().lower() if len(row) > 6 and row[6] else 'owner'
                    property_type = str(row[7]).strip().lower() if len(row) > 7 and row[7] else 'residential'
                    # 标准化可能的中文输入
                    if identity in ['房主', 'owner', '业主']:
                        identity = 'owner'
                    elif identity in ['租户', 'renter']:
                        identity = 'renter'
                    else:
                        # 保持默认
                        identity = 'owner'
                    if property_type in ['住宅', 'residential']:
                        property_type = 'residential'
                    elif property_type in ['商铺', 'commercial', '店铺']:
                        property_type = 'commercial'
                    else:
                        property_type = 'residential'

                    # 验证必填字段
                    if not room_no or not name:
                        fail_count += 1
                        errors.append(f"第{row_idx}行：房号或姓名为空")
                        continue
                    
                    # 创建住户（传入可选 building、unit、identity 和 property_type）
                    ResidentService.create_resident(
                        building=building,
                        unit=unit,
                        room_no=room_no,
                        name=name,
                        phone=phone,
                        area=area,
                        move_in_date=move_in_date,
                        identity=identity,
                        property_type=property_type
                    )
                    success_count += 1
                    
                except ValueError as e:
                    fail_count += 1
                    errors.append(f"第{row_idx}行：{str(e)}")
                except Exception as e:
                    fail_count += 1
                    errors.append(f"第{row_idx}行：导入失败 - {str(e)}")
            
            return success_count, fail_count, errors
            
        except Exception as e:
            raise Exception(f"读取Excel文件失败：{str(e)}")
    
    @staticmethod
    def create_import_template(file_path):
        """创建导入模板文件
        
        Args:
            file_path: 保存路径
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # 设置标题行（包含可选列 identity 和 property_type）
        headers = ['楼栋', '单元', '房号', '姓名', '电话', '面积', '入住日期', '身份(identity，可选：房主/租户)', '房屋类型(property_type，可选：住宅/商铺)']
        sheet.append(headers)
        
        # 设置列宽（适配新增列）
        sheet.column_dimensions['A'].width = 8   # 楼栋
        sheet.column_dimensions['B'].width = 8   # 单元
        sheet.column_dimensions['C'].width = 12  # 房号
        sheet.column_dimensions['D'].width = 15  # 姓名
        sheet.column_dimensions['E'].width = 15  # 电话
        sheet.column_dimensions['F'].width = 12  # 面积
        sheet.column_dimensions['G'].width = 15  # 入住日期
        
        # 添加示例数据（包含可选列）
        sheet.append(['6', '1', '1204', '张三', '13800138000', '80.5', '2024-01-01', '房主', '住宅'])
        sheet.append(['6', '2', '1002', '李四', '13900139000', '90.0', '2024-02-01', '租户', '商铺'])
        
        # 保存文件
        workbook.save(file_path)

